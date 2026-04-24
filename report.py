from __future__ import annotations

import os
from datetime import date
from calendar import monthrange
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import FACTORIES, ASSUMED_WORKING_DAYS, OUTPUT_DIR
from db import fetch_updates_for_month, fetch_updates_all, fetch_all_updates


THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
BOLD = Font(bold=True)


def is_off_day(factory_off_day: str, d: date) -> bool:
    return d.strftime("%A") == factory_off_day


def computed_system_target(factory_key: str, d: date) -> int:
    """
    System target computed from monthly plan and off-day rules.
    If off day -> 0 else monthly_plan / ASSUMED_WORKING_DAYS (rounded).
    """
    fc = FACTORIES[factory_key]
    if is_off_day(fc.off_day, d):
        return 0
    return int(round(fc.monthly_plan_pcs / ASSUMED_WORKING_DAYS))


def safe_pct(numer: Optional[int], denom: Optional[int]) -> Optional[float]:
    if numer is None or denom in (None, 0):
        return None
    return float(numer) / float(denom)


def generate_monthly_report_xlsx(year: int, month: int) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    headers = [
        "Date",
        "Day",
        "Prod Target (system)",
        "Daily Production",
        "Prod % vs system target",
        "Cumulative Production",
        "Lifetime Cumulative Production",
        "Daily Dispatch",
        "Cumulative Dispatch",
        "Lifetime Cumulative Dispatch",
        "Remarks",
    ]
    widths = [12, 8, 18, 16, 22, 20, 26, 16, 20, 26, 30]

    consolidated = wb.create_sheet(title="All Factories")
    consolidated["A1"] = f"All Factories — {year}-{month:02d} Daily Report"
    consolidated["A1"].font = Font(bold=True, size=14)
    consolidated.merge_cells("A1:K1")

    consolidated_headers = headers
    consolidated.append(consolidated_headers)
    for col in range(1, len(consolidated_headers) + 1):
        cell = consolidated.cell(row=2, column=col)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for i, w in enumerate(widths, start=1):
        consolidated.column_dimensions[get_column_letter(i)].width = w

    days_in_month = monthrange(year, month)[1]
    consolidated_rows: list[list[object]] = []
    overall_cum_prod = 0
    overall_cum_disp = 0

    for factory_key, fc in FACTORIES.items():
        ws = wb.create_sheet(title=fc.display_name[:31])

        # Title row
        ws["A1"] = f"{fc.display_name} — {year}-{month:02d} Daily Report"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:K1")

        ws.append(headers)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col)
            cell.font = BOLD
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER

        # Pull WhatsApp/DB updates
        updates = fetch_updates_for_month(factory_key, year, month)
        update_by_date = {u["update_date"]: u for u in updates}

        all_updates = fetch_updates_all(factory_key)
        lifetime_daily_prod: dict[str, int] = {}
        lifetime_daily_disp: dict[str, int] = {}
        for row in all_updates:
            row_date = row.get("update_date") or ""
            lifetime_daily_prod[row_date] = lifetime_daily_prod.get(row_date, 0) + int(row.get("prod_actual") or 0)
            lifetime_daily_disp[row_date] = lifetime_daily_disp.get(row_date, 0) + int(row.get("dispatch_actual") or 0)

        month_start = date(year, month, 1).isoformat()
        lifetime_cum_prod = sum(v for k, v in lifetime_daily_prod.items() if k < month_start)
        lifetime_cum_disp = sum(v for k, v in lifetime_daily_disp.items() if k < month_start)

        cum_prod = 0
        cum_disp = 0

        for day in range(1, days_in_month + 1):
            d = date(year, month, day)
            key = d.isoformat()
            u = update_by_date.get(key)

            sys_target = computed_system_target(factory_key, d)

            prod_actual = int(u["prod_actual"]) if u and u["prod_actual"] is not None else 0
            disp_actual = int(u["dispatch_actual"]) if u and u["dispatch_actual"] is not None else 0
            remarks = u["remarks"] if u else None

            cum_prod += prod_actual
            cum_disp += disp_actual

            lifetime_cum_prod += lifetime_daily_prod.get(key, 0)
            lifetime_cum_disp += lifetime_daily_disp.get(key, 0)

            prod_pct = safe_pct(prod_actual, sys_target)

            ws.append(
                [
                    d.isoformat(),
                    d.strftime("%a"),
                    sys_target,
                    prod_actual,
                    prod_pct,
                    cum_prod,
                    lifetime_cum_prod,
                    disp_actual,
                    cum_disp,
                    lifetime_cum_disp,
                    remarks,
                ]
            )

            consolidated_rows.append(
                [
                    d.isoformat(),
                    d.strftime("%a"),
                    sys_target,
                    prod_actual,
                    prod_pct,
                    cum_prod,
                    disp_actual,
                    cum_disp,
                    (f"{fc.display_name}: {remarks}" if remarks else None),
                ]
            )

        # Formatting
        ws.freeze_panes = "A3"
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        for row in ws.iter_rows(min_row=3, max_row=2 + days_in_month, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = BORDER
                if cell.column == 5 and isinstance(cell.value, float):
                    cell.number_format = "0.00%"
                if cell.column in (1, 2):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    per_day: dict[str, dict[str, object]] = {}
    for row in consolidated_rows:
        d_iso = row[0]
        if d_iso not in per_day:
            per_day[d_iso] = {
                "day": row[1],
                "prod_sys_target": 0,
                "prod_actual": 0,
                "disp_actual": 0,
                "remarks": [],
            }
        bucket = per_day[d_iso]
        bucket["prod_sys_target"] = int(bucket["prod_sys_target"]) + int(row[2])
        bucket["prod_actual"] = int(bucket["prod_actual"]) + int(row[3])
        bucket["disp_actual"] = int(bucket["disp_actual"]) + int(row[6])
        if row[8]:
            bucket["remarks"].append(row[8])

    all_updates = fetch_all_updates()
    lifetime_total_by_date_prod: dict[str, int] = {}
    lifetime_total_by_date_disp: dict[str, int] = {}
    for row in all_updates:
        row_date = row.get("update_date") or ""
        lifetime_total_by_date_prod[row_date] = lifetime_total_by_date_prod.get(row_date, 0) + int(row.get("prod_actual") or 0)
        lifetime_total_by_date_disp[row_date] = lifetime_total_by_date_disp.get(row_date, 0) + int(row.get("dispatch_actual") or 0)

    month_start = date(year, month, 1).isoformat()
    lifetime_cum_prod = sum(v for k, v in lifetime_total_by_date_prod.items() if k < month_start)
    lifetime_cum_disp = sum(v for k, v in lifetime_total_by_date_disp.items() if k < month_start)

    for day in range(1, days_in_month + 1):
        d = date(year, month, day)
        d_iso = d.isoformat()
        bucket = per_day.get(
            d_iso,
            {
                "day": d.strftime("%a"),
                "prod_sys_target": 0,
                "prod_actual": 0,
                "disp_actual": 0,
                "remarks": [],
            },
        )

        prod_actual_total = int(bucket["prod_actual"])
        disp_actual_total = int(bucket["disp_actual"])

        overall_cum_prod += prod_actual_total
        overall_cum_disp += disp_actual_total

        lifetime_cum_prod += lifetime_total_by_date_prod.get(d_iso, 0)
        lifetime_cum_disp += lifetime_total_by_date_disp.get(d_iso, 0)

        prod_pct_total = safe_pct(prod_actual_total, int(bucket["prod_sys_target"]))
        remarks_text = "; ".join(bucket["remarks"]) if bucket["remarks"] else None

        consolidated.append(
            [
                d_iso,
                bucket["day"],
                int(bucket["prod_sys_target"]),
                prod_actual_total,
                prod_pct_total,
                overall_cum_prod,
                lifetime_cum_prod,
                disp_actual_total,
                overall_cum_disp,
                lifetime_cum_disp,
                remarks_text,
            ]
        )

    consolidated.freeze_panes = "A3"
    for row in consolidated.iter_rows(
        min_row=3,
        max_row=2 + days_in_month,
        min_col=1,
        max_col=len(consolidated_headers),
    ):
        for cell in row:
            cell.border = BORDER
            if cell.column == 5 and isinstance(cell.value, float):
                cell.number_format = "0.00%"
            if cell.column in (1, 2):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    out_path = os.path.join(OUTPUT_DIR, f"factory_daily_report_{year}_{month:02d}.xlsx")
    wb.save(out_path)
    return out_path
