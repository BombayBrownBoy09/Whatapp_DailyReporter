from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook

from config import DATA_XLSX_PATH


DATA_HEADERS = [
    "factory_key",
    "update_date",
    "prod_actual",
    "prod_target",
    "dispatch_actual",
    "dispatch_target",
    "remarks",
    "raw_message",
    "sender_phone",
    "created_at",
    "email_sent_at",
    "email_payload_hash",
]


def _data_path() -> Path:
    return Path(DATA_XLSX_PATH)


def init_db() -> None:
    path = _data_path()
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        updated = False
        if "email_sent_at" not in headers:
            ws.cell(row=1, column=len(headers) + 1, value="email_sent_at")
            headers.append("email_sent_at")
            updated = True
        if "email_payload_hash" not in headers:
            ws.cell(row=1, column=len(headers) + 1, value="email_payload_hash")
            updated = True
        if updated:
            wb.save(path)
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "raw_data"
    ws.append(DATA_HEADERS)
    wb.save(path)


def upsert_daily_update(
    *,
    factory_key: str,
    update_date: date,
    prod_actual: Optional[int],
    prod_target: Optional[int],
    dispatch_actual: Optional[int],
    dispatch_target: Optional[int],
    remarks: Optional[str],
    raw_message: str,
    sender_phone: str,
) -> bool:
    init_db()
    path = _data_path()
    wb = load_workbook(path)
    ws = wb["raw_data"]

    target_date = update_date.isoformat()
    found_row = None
    for row in ws.iter_rows(min_row=2, values_only=False):
        if (row[0].value == factory_key) and (row[1].value == target_date):
            found_row = row
            break

    existing_email_sent_at = None
    existing_email_hash = None
    if found_row:
        if len(found_row) >= len(DATA_HEADERS) - 1:
            existing_email_sent_at = found_row[len(DATA_HEADERS) - 2].value
            existing_email_hash = found_row[len(DATA_HEADERS) - 1].value

    values = [
        factory_key,
        target_date,
        prod_actual,
        prod_target,
        dispatch_actual,
        dispatch_target,
        remarks,
        raw_message,
        sender_phone,
        date.today().isoformat(),
        existing_email_sent_at,
        existing_email_hash,
    ]

    inserted = False
    if found_row:
        for idx, cell in enumerate(found_row):
            cell.value = values[idx]
    else:
        ws.append(values)
        inserted = True

    wb.save(path)
    return inserted


def _row_to_dict(row: tuple[Any, ...]) -> dict[str, Any]:
    padded = list(row)
    if len(padded) < len(DATA_HEADERS):
        padded.extend([None] * (len(DATA_HEADERS) - len(padded)))
    return dict(zip(DATA_HEADERS, padded))


def mark_email_sent(
    factory_key: str,
    update_date: date,
    *,
    sent_at: str | None = None,
    payload_hash: str | None = None,
) -> None:
    init_db()
    path = _data_path()
    wb = load_workbook(path)
    ws = wb["raw_data"]
    target_date = update_date.isoformat()
    sent_at_value = sent_at or date.today().isoformat()

    for row in ws.iter_rows(min_row=2, values_only=False):
        if (row[0].value == factory_key) and (row[1].value == target_date):
            row[len(DATA_HEADERS) - 2].value = sent_at_value
            row[len(DATA_HEADERS) - 1].value = payload_hash
            wb.save(path)
            return


def is_email_up_to_date(factory_key: str, update_date: date, payload_hash: str) -> bool:
    record = fetch_daily_update(factory_key, update_date)
    if not record:
        return False
    return bool(record.get("email_payload_hash") == payload_hash)


def fetch_daily_update(factory_key: str, update_date: date) -> Optional[dict[str, Any]]:
    init_db()
    path = _data_path()
    if not path.exists():
        return None

    wb = load_workbook(path, read_only=True)
    ws = wb["raw_data"]
    target_date = update_date.isoformat()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == factory_key and row[1] == target_date:
            return _row_to_dict(row)
    return None


def fetch_updates_for_month(factory_key: str, year: int, month: int) -> list[dict[str, Any]]:
    start = date(year, month, 1).isoformat()
    # next month:
    if month == 12:
        end = date(year + 1, 1, 1).isoformat()
    else:
        end = date(year, month + 1, 1).isoformat()

    init_db()
    path = _data_path()
    if not path.exists():
        return []

    wb = load_workbook(path, read_only=True)
    ws = wb["raw_data"]
    results: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_date = row[1] or ""
        if row[0] != factory_key:
            continue
        if row_date < start or row_date >= end:
            continue
        results.append(_row_to_dict(row))
    results.sort(key=lambda r: r.get("update_date") or "")
    return results


def fetch_updates_all(factory_key: str) -> list[dict[str, Any]]:
    init_db()
    path = _data_path()
    if not path.exists():
        return []

    wb = load_workbook(path, read_only=True)
    ws = wb["raw_data"]
    results: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] != factory_key:
            continue
        results.append(_row_to_dict(row))
    results.sort(key=lambda r: r.get("update_date") or "")
    return results


def fetch_all_updates() -> list[dict[str, Any]]:
    init_db()
    path = _data_path()
    if not path.exists():
        return []

    wb = load_workbook(path, read_only=True)
    ws = wb["raw_data"]
    results: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        results.append(_row_to_dict(row))
    results.sort(key=lambda r: r.get("update_date") or "")
    return results
